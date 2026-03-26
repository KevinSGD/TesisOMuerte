import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def export_excel(df_asig, df_cal, df_uso, solution, data, output_dir=None):
    base = output_dir or "outputs"
    os.makedirs(base, exist_ok=True)
    out = os.path.join(base, f"horario_profesores_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx")

    meta = solution["meta"]

    # Estadísticas
    df_stats = pd.DataFrame([
        {"Métrica": "Estado", "Valor": str(solution["status"])},
        {"Métrica": "Tiempo ejecución (s)", "Valor": solution["elapsed"]},
        {"Métrica": "Variables", "Valor": solution["num_vars"]},
        {"Métrica": "Restricciones", "Valor": solution["num_constraints"]},
        {"Métrica": "Total clases (bloques)", "Valor": len(df_asig)},
        {"Métrica": "Total grupos", "Valor": meta["num_grupos"]},
        {"Métrica": "Cursos con 3 grupos (aleatorios disciplinares)", "Valor": len(meta["cursos_con_3"])},
        {"Métrica": "Podas por tipo/capacidad", "Valor": meta["podas"]},
        {"Métrica": "Checks tipo/capacidad", "Valor": meta["total_checks"]},
        {"Métrica": "Profesor Infra (índice)", "Valor": meta["p_infra"]},
        {"Métrica": "Salón Infra (índice)", "Valor": meta["s_infra"] if meta["s_infra"] is not None else ""},
    ])

    # Capacidad salones
    salones = meta["salones"]
    tipo_salon = meta["tipo_salon"]
    capacidad = meta["capacidad_salon"]

    df_cap = pd.DataFrame([
        {"Salón": salones[s], "Tipo": tipo_salon[s], "Capacidad": capacidad[s]}
        for s in range(len(salones))
    ]).sort_values(["Tipo", "Salón"])

    # Clasificación materias
    materias = meta["materias"]
    areas = meta["areas"]
    creditos = meta["creditos"]
    tipo_materia = meta["tipo_materia"]
    df_clas = pd.DataFrame([
        {"Materia": materias[c], "Área": areas[c], "Créditos": creditos[c], "Tipo": tipo_materia[c]}
        for c in range(len(materias))
    ]).sort_values(["Tipo", "Área", "Materia"])

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Formato como antes
        if df_asig.empty:
            pd.DataFrame(columns=["Día", "Bloque", "Curso", "Materia", "Profesor", "Salón"]).to_excel(
                writer, sheet_name="Asignaciones", index=False
            )
        else:
            df_asig[["Día", "Bloque", "Curso", "Materia", "Profesor", "Salón"]].to_excel(
                writer, sheet_name="Asignaciones", index=False
            )

        df_cal.to_excel(writer, sheet_name="Calendario_Visual", index=False)

        if df_uso.empty:
            pd.DataFrame(columns=["Día", "Bloque", "Salón", "Curso", "Profesor"]).to_excel(
                writer, sheet_name="Uso_Salones", index=False
            )
        else:
            df_uso[["Día", "Bloque", "Salón", "Curso", "Profesor"]].to_excel(
                writer, sheet_name="Uso_Salones", index=False
            )

        df_stats.to_excel(writer, sheet_name="Estadisticas", index=False)
        df_cap.to_excel(writer, sheet_name="Capacidad_Salones", index=False)
        df_clas.to_excel(writer, sheet_name="Clasificacion_Materias", index=False)

    # Ajuste visual hoja Calendario_Visual (como antes)
    wb = load_workbook(out)
    ws = wb["Calendario_Visual"]

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    col_widths = {"A": 10, "B": 38, "C": 38, "D": 38, "E": 38, "F": 38}
    for col, w in col_widths.items():
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = w

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 95

    wb.save(out)
    return out